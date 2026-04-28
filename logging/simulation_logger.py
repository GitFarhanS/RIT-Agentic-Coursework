"""
Log each tick of the RIT simulation to an xlsx file.
- One row per tick: best bid/ask for each security, tender info.
- One sheet per ticker: full order book depth (20 levels) per tick.
- One sheet: all tenders seen across the run.
One xlsx file is produced per simulation run (in simulation_logs/).
Tenders are observed and logged only — never accepted or declined.
"""

import time
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from sdk import RotmanSDK

# Config

API_KEY = "TPIAOJIF"
HOST = "http://192.168.64.9:9999/v1"
POLL_INTERVAL_SEC = 1.0
OUTPUT_DIR = Path(__file__).resolve().parent / "simulation_logs"
BOOK_DEPTH = 20

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
CENTER = Alignment(horizontal="center")

STOPPED_WAIT_TIMEOUT = 800  # seconds to wait for ACTIVE before giving up
MAX_SESSIONS = None  # set to None to loop forever



# Helpers

def _wait_for_session_start(client: RotmanSDK) -> bool:
    """Wait for a true new session start: STOPPED → ACTIVE @ tick 0 transition."""
    print("Waiting for session start...", flush=True)
    deadline = time.monotonic() + STOPPED_WAIT_TIMEOUT

    # Step 1: If already ACTIVE, wait for it to go STOPPED first
    # (so we catch the next clean start, not mid-session)
    case = client.get_case()
    status = case.get("status", "")
    if status == "ACTIVE":
        print("  Sim is mid-session — waiting for it to STOP first...", flush=True)
        while time.monotonic() < deadline:
            case = client.get_case()
            if case.get("status") == "STOPPED":
                print("  STOPPED — now waiting for next ACTIVE...", flush=True)
                break
            time.sleep(1.0)
        else:
            return False

    # Step 2: Wait for ACTIVE @ tick 0 — true session start
    while time.monotonic() < deadline:
        case = client.get_case()
        status = case.get("status", "")
        tick = int(case.get("tick", 0) or 0)
        if status == "ACTIVE" and tick == 0:
            print(f"  New session started at tick 0.", flush=True)
            return True
        time.sleep(0.5)

    return False

def _style_header_row(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _style_data_rows(ws, n_cols: int) -> None:
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            if fill:
                cell.fill = fill


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 8), 30)


def _bid_ask(sec: dict) -> tuple[str, float, float]:
    ticker = sec.get("ticker", "")
    bid = float(sec.get("bid") or sec.get("bidPrice") or 0)
    ask = float(sec.get("ask") or sec.get("askPrice") or 0)
    return ticker, bid, ask


def _fetch_book(client: RotmanSDK, ticker: str) -> tuple[list, list]:
    try:
        book = client.get_securities_book(ticker, limit=BOOK_DEPTH)
        bids = [(lev.price, getattr(lev, "quantity_remaining", 0)) for lev in book.bids[:BOOK_DEPTH]]
        asks = [(lev.price, getattr(lev, "quantity_remaining", 0)) for lev in book.asks[:BOOK_DEPTH]]
        return bids, asks
    except Exception as e:
        print(f"  [warn] get_securities_book({ticker}) failed: {e}", flush=True)
        return [], []


def _fetch_tenders(client: RotmanSDK) -> list[dict]:
    try:
        raw = client.get_tenders()
    except Exception:
        return []
    if not raw:
        return []
    result = []
    for t in raw:
        action = getattr(t, "action", None)
        result.append({
            "tender_id": getattr(t, "id", ""),
            "ticker":    getattr(t, "ticker", ""),
            "action":    str(action) if action is not None else "",
            "quantity":  getattr(t, "quantity", 0),
            "price":     getattr(t, "price", 0),
            "expires":   getattr(t, "expiry_tick", 0),
        })
    return result



# Main logger

def run_simulation_logger() -> None:
    client = RotmanSDK(API_KEY=API_KEY, HOST=HOST)
    session = 0

    while MAX_SESSIONS is None or session < MAX_SESSIONS:
        session += 1
        print(f"\n{'='*50}", flush=True)
        print(f"Session {session}{f'/{MAX_SESSIONS}' if MAX_SESSIONS else ''}", flush=True)
        print(f"{'='*50}", flush=True)

        if not _wait_for_session_start(client):
            print(f"Timed out waiting for session {session} start. Stopping.")
            break

        case = client.get_case()
        print(f"Case: {case.get('name', 'unknown')} | status: {case.get('status')} | tick: {case.get('tick')}", flush=True)

        snapshots = _collect_session(client)

        if not snapshots:
            print("No data collected for this session.")
            continue

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = OUTPUT_DIR / f"simulation_run_s{session:02d}_{run_id}.xlsx"
        _write_xlsx(snapshots, out_path)
        print(f"Saved: {out_path}", flush=True)

    print(f"\nDone. Logged {session} session(s).")

def _collect_session(client: RotmanSDK) -> list[dict]:
    """Collect all tick snapshots for a single session. Returns list of snapshots."""
    snapshots: list[dict] = []
    seen_ticks: set[int] = set()
    prev_bid_ask: dict[str, tuple[float, float]] = {}
    prev_books: dict[str, tuple[list, list]] = {}

    try:
        while True:
            case = client.get_case()
            status = case.get("status", "")
            tick = int(case.get("tick", 0) or 0)

            if status == "STOPPED":
                if snapshots:
                    print(f"\nSession ended at tick {max(seen_ticks)}.", flush=True)
                    break
                time.sleep(POLL_INTERVAL_SEC)
                continue

            if tick in seen_ticks:
                time.sleep(POLL_INTERVAL_SEC)
                continue

            seen_ticks.add(tick)

            securities_raw = client.get_securities()
            if not isinstance(securities_raw, list):
                securities_raw = [securities_raw] if securities_raw else []

            stocks = []
            for sec in securities_raw:
                ticker, bid, ask = _bid_ask(sec)
                if not ticker:
                    continue
                if prev_bid_ask.get(ticker) != (bid, ask):
                    bids, asks = _fetch_book(client, ticker)
                    prev_books[ticker] = (bids, asks)
                    prev_bid_ask[ticker] = (bid, ask)
                else:
                    bids, asks = prev_books.get(ticker, ([], []))
                stocks.append({"ticker": ticker, "bid": bid, "ask": ask, "bids": bids, "asks": asks})

            tenders = _fetch_tenders(client)
            snapshots.append({"tick": tick, "stocks": stocks, "tenders": tenders})

            tender_info = f"{len(tenders)} tender(s)" if tenders else "no tenders"
            print(f"  tick {tick:>3} | {len(stocks)} securities | {tender_info}", flush=True)

            if tick >= 299:
                print("Final tick reached.", flush=True)
                break

            time.sleep(POLL_INTERVAL_SEC)

    except KeyboardInterrupt:
        print("\nInterrupted — saving collected data.", flush=True)
        raise  # re-raise so the outer loop also exits cleanly

    return snapshots


def _write_xlsx(snapshots: list[dict], out_path: Path) -> None:
    all_tickers: list[str] = []
    for snap in snapshots:
        for st in snap["stocks"]:
            t = st["ticker"]
            if t and t not in all_tickers:
                all_tickers.append(t)

    wb = Workbook()

    # Sheet 1: "By Tick" — one row per tick, best bid/ask + tender info
    ws = wb.active
    ws.title = "By Tick"

    headers = ["Tick"]
    for t in all_tickers:
        headers += [f"{t} Bid", f"{t} Ask", f"{t} Spread"]
    headers += ["Has Tender", "Tender Ticker", "Tender Action", "Tender Qty", "Tender Price", "Tender Expires"]
    ws.append(headers)

    for snap in snapshots:
        row = [snap["tick"]]
        ba = {st["ticker"]: (st["bid"], st["ask"]) for st in snap["stocks"]}
        for t in all_tickers:
            bid, ask = ba.get(t, (None, None))
            spread = round(ask - bid, 4) if bid and ask else ""
            row += [bid or "", ask or "", spread]
        t0 = snap["tenders"][0] if snap["tenders"] else None
        if t0:
            row += ["Y", t0["ticker"], t0["action"], t0["quantity"], t0["price"], t0["expires"]]
        else:
            row += ["N", "", "", "", "", ""]
        ws.append(row)

    _style_header_row(ws, len(headers))
    _style_data_rows(ws, len(headers))
    _auto_width(ws)
    ws.freeze_panes = "A2"

    # Sheet per ticker: full order book depth per tick
    book_headers = ["Tick"]
    for i in range(1, BOOK_DEPTH + 1):
        book_headers += [f"Bid {i} Price", f"Bid {i} Qty"]
    for i in range(1, BOOK_DEPTH + 1):
        book_headers += [f"Ask {i} Price", f"Ask {i} Qty"]

    tick_to_stocks = {snap["tick"]: {st["ticker"]: st for st in snap["stocks"]} for snap in snapshots}

    for ticker in all_tickers:
        ws_b = wb.create_sheet(f"Book {ticker}"[:31])
        ws_b.append(book_headers)
        for snap in snapshots:
            st = tick_to_stocks.get(snap["tick"], {}).get(ticker)
            bids = (st["bids"] if st else [])[:BOOK_DEPTH]
            asks = (st["asks"] if st else [])[:BOOK_DEPTH]
            row = [snap["tick"]]
            for i in range(BOOK_DEPTH):
                row += [bids[i][0], bids[i][1]] if i < len(bids) else ["", ""]
            for i in range(BOOK_DEPTH):
                row += [asks[i][0], asks[i][1]] if i < len(asks) else ["", ""]
            ws_b.append(row)
        _style_header_row(ws_b, len(book_headers))
        _style_data_rows(ws_b, len(book_headers))
        ws_b.freeze_panes = "B2"

    # Sheet: Tenders log
    all_tenders = [(snap["tick"], t) for snap in snapshots for t in snap["tenders"]]
    if all_tenders:
        ws_t = wb.create_sheet("Tenders")
        t_headers = ["Tick", "Tender ID", "Ticker", "Action", "Quantity", "Price", "Expires"]
        ws_t.append(t_headers)
        for tick, t in all_tenders:
            ws_t.append([tick, t["tender_id"], t["ticker"], t["action"], t["quantity"], t["price"], t["expires"]])
        _style_header_row(ws_t, len(t_headers))
        _style_data_rows(ws_t, len(t_headers))
        _auto_width(ws_t)
        ws_t.freeze_panes = "A2"

    wb.save(out_path)
    wb.close()


if __name__ == "__main__":
    run_simulation_logger()