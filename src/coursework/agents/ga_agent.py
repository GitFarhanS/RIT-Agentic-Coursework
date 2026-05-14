"""
GA-driven Deliberative Agent
----------------------------
Applies evolved parameters via ``DeliberativeSettings.from_ga_params`` (no module
mutation) and adds an edge-threshold pre-screen on tender evaluation.
"""

from __future__ import annotations

import argparse
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from coursework.agents import deliberative as deliberative_pkg
from coursework.agents.common import get_mid_price, style_worksheet, wait_for_session_start
from coursework.agents.deliberative.agent import DeliberativeAgent
from coursework.agents.deliberative.records import SessionStats, TenderRecord
from coursework.agents.deliberative.settings import DeliberativeSettings
from coursework.config.runtime import load_runtime_config
from coursework.infrastructure.rotman_client import RITError, RotmanSDK


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[3]


_EVOLVED_PATH = _repo_root() / "ga" / "evolved_params.json"
EVOLVED = json.loads(_EVOLVED_PATH.read_text())
PARAMS = EVOLVED["params"]

MARKET_ORDER_RATIO = float(PARAMS["market_order_ratio"])
URGENCY_THRESHOLD = int(300 * float(PARAMS["time_decay_factor"]) / 3.0)
SLICE_SIZE_RATIO = float(PARAMS["slice_size_ratio"])
EDGE_THRESHOLD = float(PARAMS["threshold"])

# Tests monkeypatch ``ga.base.DeliberativeAgent`` — keep alias to deliberative package
base = deliberative_pkg


class GAAgent(DeliberativeAgent):
    """Deliberative agent + GA edge gate; unwind slice sizing comes from injected settings."""

    def __init__(
        self,
        client: RotmanSDK,
        settings: DeliberativeSettings | None = None,
        *,
        edge_threshold: float | None = None,
    ) -> None:
        s = settings or DeliberativeSettings.from_ga_params(PARAMS)
        super().__init__(client, s)
        self._edge_threshold = float(edge_threshold if edge_threshold is not None else PARAMS["threshold"])

    def evaluate(self, tender: Any, tick: int) -> TenderRecord | None:
        tid = getattr(tender, "id", None)
        ticker = getattr(tender, "ticker", None)
        price = float(getattr(tender, "price", 0) or 0)
        qty = int(getattr(tender, "quantity", 0) or 0)
        action = str(getattr(tender, "action", "") or "").upper()

        if tid and ticker and price > 0:
            mid = get_mid_price(self.client, ticker)
            if mid:
                edge_signed = (mid - price) / mid
                edge = abs(edge_signed)
                if edge < self._edge_threshold:
                    try:
                        self.client.decline_tender(int(tid))
                    except RITError as e:
                        print(f"  [warn] decline {tid}: {e}", flush=True)
                    print(
                        f"  [DECLINED] tid={tid} edge={edge:.4f} < thr={self._edge_threshold:.4f}",
                        flush=True,
                    )
                    return TenderRecord(
                        tick=tick,
                        tender_id=int(tid),
                        ticker=str(ticker),
                        action=action,
                        quantity=qty,
                        tender_price=price,
                        mid_price=mid,
                        edge=edge,
                        edge_signed=edge_signed,
                        decision="DECLINED",
                        decline_reason=f"edge<{self._edge_threshold:.4f}",
                    )
        return super().evaluate(tender, tick)


def _write_xlsx(all_stats: list[SessionStats], out_path: Path) -> None:
    wb = Workbook()
    params_str = json.dumps(PARAMS, sort_keys=True)

    ws = wb.active
    ws.title = "Summary"
    headers = ["Session", "Seen", "Accepted", "Declined", "Accept %", "Gross P&L", "Net P&L", "Params"]
    ws.append(headers)
    for s in all_stats:
        rate = f"{s.accepted / s.seen:.1%}" if s.seen else "0.0%"
        ws.append(
            [
                s.session,
                s.seen,
                s.accepted,
                s.declined,
                rate,
                round(s.total_gross_pnl, 4),
                round(s.total_net_pnl, 4),
                params_str,
            ]
        )
    if len(all_stats) > 1:
        ts = sum(s.seen for s in all_stats)
        ta = sum(s.accepted for s in all_stats)
        td = sum(s.declined for s in all_stats)
        ws.append(
            [
                "TOTAL",
                ts,
                ta,
                td,
                f"{ta/ts:.1%}" if ts else "0.0%",
                round(sum(s.total_gross_pnl for s in all_stats), 4),
                round(sum(s.total_net_pnl for s in all_stats), 4),
                params_str,
            ]
        )
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
    style_worksheet(ws, n_cols=len(headers))

    ws2 = wb.create_sheet("Tender Detail")
    rec_headers = [
        "Session",
        "Tick",
        "Tender ID",
        "Ticker",
        "Action",
        "Quantity",
        "Tender Price",
        "Mid Price",
        "Edge",
        "Edge Signed",
        "Est Slippage",
        "Spread Captured",
        "Txn Cost",
        "Expected PnL",
        "Risk Blocked",
        "Decision",
        "Decline Reason",
        "Avg Unwind Price",
        "Gross P&L",
        "Commission",
        "Liab Violation Vol",
        "Liab Fine",
        "Net P&L",
        "Residual",
        "Params",
    ]
    ws2.append(rec_headers)
    for s in all_stats:
        for r in s.records:
            ws2.append(
                [
                    s.session,
                    r.tick,
                    r.tender_id,
                    r.ticker,
                    r.action,
                    r.quantity,
                    r.tender_price,
                    r.mid_price,
                    r.edge,
                    r.edge_signed,
                    r.estimated_slippage,
                    r.spread_captured,
                    r.txn_cost,
                    r.expected_pnl,
                    r.risk_blocked,
                    r.decision,
                    r.decline_reason,
                    r.avg_unwind_price,
                    r.gross_pnl,
                    r.commission_cost,
                    r.liability_violation_volume,
                    r.liability_fine,
                    r.net_pnl,
                    r.residual,
                    params_str,
                ]
            )
    style_worksheet(ws2, n_cols=len(rec_headers))
    ws2.freeze_panes = "A2"

    green = PatternFill("solid", start_color="C6EFCE")
    red = PatternFill("solid", start_color="FFC7CE")
    col = rec_headers.index("Decision") + 1
    for row in ws2.iter_rows(min_row=2, min_col=col, max_col=col):
        for c in row:
            if c.value == "ACCEPTED":
                c.fill = green
            elif c.value == "DECLINED":
                c.fill = red

    wb.save(out_path)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="GA-driven deliberative LT3 agent")
    parser.add_argument(
        "--match-anchors",
        type=Path,
        metavar="DIR",
        default=None,
        help="Before each session, wait until tick 0 matches DIR/r_<session>.json from a reactive run",
    )
    args = parser.parse_args()

    print(f"[GA] params: {PARAMS}", flush=True)
    runtime = load_runtime_config()
    client = RotmanSDK(API_KEY=runtime.api_key, HOST=runtime.host)
    agent = GAAgent(client)
    st = agent.settings
    all_stats: list[SessionStats] = []
    session = 0
    match_timeout = float(os.environ.get("RIT_ANCHOR_MATCH_TIMEOUT", "1200"))

    try:
        while st.max_sessions is None or session < st.max_sessions:
            session += 1
            print(
                f"\n{'=' * 60}\n  SESSION {session}/{st.max_sessions} | "
                f"mkt={MARKET_ORDER_RATIO} urgency<{URGENCY_THRESHOLD} "
                f"slice={SLICE_SIZE_RATIO} edge_thr={EDGE_THRESHOLD}\n{'=' * 60}",
                flush=True,
            )
            if args.match_anchors is not None:
                from coursework.agents.session_pairing import load_fingerprint, wait_for_matched_session_start

                anchor_path = args.match_anchors / f"r_{session}.json"
                if not anchor_path.is_file():
                    print(f"[pairing] missing anchor file {anchor_path} — run reactive with --save-anchors first.")
                    break
                target = load_fingerprint(anchor_path)
                if not wait_for_matched_session_start(
                    client,
                    target,
                    stopped_wait_timeout=st.stopped_timeout,
                    match_timeout_sec=match_timeout,
                ):
                    print(f"Timed out waiting for matching scenario for session {session}.")
                    break
            elif not wait_for_session_start(client=client, stopped_wait_timeout=st.stopped_timeout):
                print(f"Timed out waiting for session {session}.")
                break
            stats = agent.run_session(session)
            all_stats.append(stats)
            print(
                f"\n[session {session}] seen={stats.seen} acc={stats.accepted} "
                f"dec={stats.declined} gross={stats.total_gross_pnl:+.2f} "
                f"net={stats.total_net_pnl:+.2f}",
                flush=True,
            )
    except KeyboardInterrupt:
        print("\nInterrupted — saving logs.", flush=True)

    if not all_stats:
        print("No sessions completed.")
        return

    st.output_dir.mkdir(parents=True, exist_ok=True)
    out = st.output_dir / f"ga_agent_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    _write_xlsx(all_stats, out)
    print(f"\nSaved: {out}", flush=True)


if __name__ == "__main__":
    main()
