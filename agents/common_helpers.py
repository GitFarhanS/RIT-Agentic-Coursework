from __future__ import annotations

import time
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill


def style_worksheet(ws: Any, n_cols: int, max_col_width: int = 35) -> None:
    """Apply shared worksheet styling for agent logs."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = PatternFill("solid", start_color="D6E4F0") if row_idx % 2 == 0 else None
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            if fill:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), max_col_width)


def get_mid_price(client: Any, ticker: str) -> float | None:
    """Return mid-price for ticker from current securities snapshot."""
    try:
        secs = client.get_securities(ticker=ticker)
        if not secs:
            return None
        sec = secs[0] if isinstance(secs, list) else secs
        bid = float(sec.get("bid") or sec.get("bidPrice") or 0)
        ask = float(sec.get("ask") or sec.get("askPrice") or 0)
        if bid > 0 and ask > 0:
            return (bid + ask) / 2
        return None
    except Exception as e:
        print(f"  [warn] _get_mid({ticker}): {e}", flush=True)
        return None


def get_last_trade_price(client: Any, ticker: str) -> float:
    """Best-effort last traded price from securities endpoint."""
    try:
        secs = client.get_securities(ticker=ticker)
        if not secs:
            return 0.0
        sec = secs[0] if isinstance(secs, list) else secs
        return float(sec.get("last") or sec.get("last_price") or sec.get("lastPrice") or 0.0)
    except Exception as e:
        print(f"    [warn] _get_last_trade_price({ticker}): {e}", flush=True)
        return 0.0


def wait_for_session_start(client: Any, stopped_wait_timeout: float) -> bool:
    """Wait for STOPPED->ACTIVE(tick 0) boundary and return True if detected."""
    print("Waiting for session start...", flush=True)
    deadline = time.monotonic() + stopped_wait_timeout

    case = client.get_case()
    if case.get("status") == "ACTIVE":
        print("  Sim is mid-session — waiting for STOP first...", flush=True)
        while time.monotonic() < deadline:
            if client.get_case().get("status") == "STOPPED":
                print("  STOPPED — waiting for next ACTIVE...", flush=True)
                break
            time.sleep(1.0)
        else:
            return False

    while time.monotonic() < deadline:
        case = client.get_case()
        status = case.get("status", "")
        tick = int(case.get("tick", 0) or 0)
        if status == "ACTIVE" and tick == 0:
            print("  New session detected at tick 0.", flush=True)
            return True
        time.sleep(0.5)

    return False
