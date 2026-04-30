"""
Phase 2 — Deliberative Agent
----------------------------
Tenders: order-book slippage model, spread vs tender price, txn cost, exposure limits.
Accept only if projected expected PnL > 0 and within net/gross headroom.
Unwind: market chunk + passive limits, repricing every ~10 ticks, urgency near session end.
"""

from __future__ import annotations

import math
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from common_helpers import get_last_trade_price, get_mid_price, style_worksheet, wait_for_session_start
from runtime_config import load_runtime_config
from sdk import RITError, RotmanSDK
from utilities import ActionEnum, OrderType


# --- Config (GA will replace constants later)

_RUNTIME = load_runtime_config()
API_KEY = _RUNTIME.api_key
HOST = _RUNTIME.host

POLL_INTERVAL_SEC = 0.5
MAX_SESSIONS = 3
STOPPED_WAIT_TIMEOUT = 700

OUTPUT_DIR = Path(__file__).resolve().parent / "agent_logs"

MAX_ORDER_QTY: dict[str, int] = {
    "CRZY": 25_000,
    "TAME": 10_000,
}
DEFAULT_MAX_QTY = 10_000

BOOK_DEPTH = 20
NO_TRADE_TICKS_BEFORE_EXPIRY = 0  # only decline tenders at/after expiry

MARKET_ORDER_RATIO = 0.6
TIME_DECAY_THRESHOLD = 60  # urgency if fewer than this many session ticks remaining

NET_HEADROOM_CEILING = 100_000  # projected |portfolio net|
GROSS_HEADROOM_CEILING = 250_000  # projected gross exposure
COMMISSION_PER_SHARE = 0.02
LIABILITY_FINE_PER_SHARE = 0.39


# --- Dataclasses


@dataclass
class TenderRecord:
    tick: int
    tender_id: int
    ticker: str
    action: str
    quantity: int
    tender_price: float
    mid_price: float
    edge: float
    decision: str
    bid_depth_total: int = 0
    ask_depth_total: int = 0
    n_bid_levels: int = 0
    best_bid: float = 0.0
    best_ask: float = 0.0
    estimated_slippage: float = 0.0
    book_covers_full_qty: bool = False
    unwind_method: str = ""
    spread_captured: float = 0.0
    slippage_est: float = 0.0
    txn_cost: float = 0.0
    expected_pnl: float = 0.0
    risk_blocked: bool = False
    fill_price: float = 0.0
    pnl: float = 0.0  # gross trading P&L
    commission_cost: float = 0.0
    liability_violation_volume: int = 0
    liability_fine: float = 0.0
    net_pnl: float = 0.0
    residual: int = 0


@dataclass
class SessionStats:
    session: int
    tenders_seen: int = 0
    tenders_accepted: int = 0
    tenders_declined: int = 0
    total_pnl: float = 0.0  # net P&L
    total_gross_pnl: float = 0.0
    records: list[TenderRecord] = field(default_factory=list)


@dataclass
class UnwindState:
    """Live unwind after a tender accept."""

    tender_id: int
    ticker: str
    quantity: int
    tender_price: float
    tender_was_buy: bool
    unwind_action: ActionEnum
    limit_order_ids: list[int] = field(default_factory=list)
    last_mgmt_tick: int = -1
    done: bool = False
    # Realized from market executions only; missing passive fill volume filled with mid proxy at PnL time
    unwind_value_accum: float = 0.0
    unwind_qty_accum: int = 0
    used_limit_orders: bool = False
    liability_remaining: int = 0


@dataclass(frozen=True)
class TenderInput:
    tender_id: int
    ticker: str
    quantity: int
    tender_price: float
    expiry_tick: int
    action_str: str
    tender_is_buy: bool


@dataclass(frozen=True)
class TenderMarketSnapshot:
    mid: float
    bid_depth_total: int
    ask_depth_total: int
    n_bid_levels: int
    best_bid: float
    best_ask: float
    slip: float
    book_covers_full_qty: bool


@dataclass(frozen=True)
class TenderDecision:
    edge: float
    spread_captured: float
    txn_cost: float
    expected_pnl: float
    risk_blocked: bool
    liability_safe: bool
    decline_reason: str | None


# --- Styling helpers (same spirit as reactive)


def _style_ws(ws: Any, n_cols: int) -> None:
    style_worksheet(ws, n_cols=n_cols, max_col_width=38)


class DeliberativeAgent:
    def __init__(self, client: RotmanSDK) -> None:
        self.client = client
        self._unwind: UnwindState | None = None

    # ---------- Market data ----------

    def _get_mid(self, ticker: str) -> float | None:
        return get_mid_price(self.client, ticker)

    def _get_last_trade_price(self, ticker: str) -> float:
        return get_last_trade_price(self.client, ticker)

    def _best_bb(self, ticker: str) -> tuple[float, float]:
        """Top-of-book bid and ask."""
        secs = self.client.get_securities(ticker=ticker)
        if not secs:
            return 0.0, 0.0
        sec = secs[0] if isinstance(secs, list) else secs
        bid = float(sec.get("bid") or sec.get("bidPrice") or 0)
        ask = float(sec.get("ask") or sec.get("askPrice") or 0)
        return bid, ask

    @staticmethod
    def _infer_tick(best_bid: float, best_ask: float) -> float:
        _ = (best_bid, best_ask)
        return 0.01

    @staticmethod
    def _limit_sell_inside_spread(best_bid: float, best_ask: float, tick_sz: float) -> float:
        px = round(best_ask - tick_sz, 4)
        if best_bid > 0:
            px = max(px, best_bid)
        return max(px, tick_sz)

    @staticmethod
    def _limit_buy_inside_spread(best_bid: float, best_ask: float, tick_sz: float) -> float:
        px = round(best_bid + tick_sz, 4)
        if best_ask > 0:
            px = min(px, best_ask)
        return px

    # ---------- Slippage ----------

    @staticmethod
    def _level_avail(level: Any) -> int:
        return int(level.quantity_remaining or level.quantity or 0)

    def _estimate_slippage_buy_tender(
        self,
        bids: list[Any],
        walk_qty: int,
        best_bid_px: float,
    ) -> tuple[float, bool]:
        """Walk bid side sell (unwind BUY tender). Sum (best_bid - level_px) * q."""
        if walk_qty <= 0:
            return 0.0, True
        if not bids or best_bid_px <= 0:
            return float("inf"), False
        remaining = walk_qty
        slip = 0.0
        for lvl in bids:
            px = float(lvl.price)
            av = self._level_avail(lvl)
            if av <= 0:
                continue
            take = min(remaining, av)
            slip += (best_bid_px - px) * take
            remaining -= take
            if remaining <= 0:
                return slip, True
        return slip, False

    def _estimate_slippage_sell_tender(
        self,
        asks: list[Any],
        walk_qty: int,
        best_ask_px: float,
    ) -> tuple[float, bool]:
        """Walk ask side buy (unwind SELL tender). Sum (level_px - best_ask) * q."""
        if walk_qty <= 0:
            return 0.0, True
        if not asks or best_ask_px <= 0:
            return float("inf"), False
        remaining = walk_qty
        slip = 0.0
        for lvl in asks:
            px = float(lvl.price)
            av = self._level_avail(lvl)
            if av <= 0:
                continue
            take = min(remaining, av)
            slip += (px - best_ask_px) * take
            remaining -= take
            if remaining <= 0:
                return slip, True
        return slip, False

    # ---------- Risk ----------

    def _projected_exposure_breach(self, ticker: str, quantity: int, tender_is_buy: bool) -> bool:
        """
        True if accepting would breach net or gross headroom (100k / 250k).
        Projected positions: current + delta from tender fill.
        """
        try:
            positions = self.client.get_positions()
        except Exception as e:
            print(f"  [warn] get_positions failed in risk check: {e}", flush=True)
            return True

        proj: dict[str, int] = {t: int(p.position) for t, p in positions.items()}
        if tender_is_buy:
            proj[ticker] = proj.get(ticker, 0) + quantity
        else:
            proj[ticker] = proj.get(ticker, 0) - quantity

        gross = sum(abs(v) for v in proj.values())
        net_mag = abs(sum(proj.values()))

        breached = gross > GROSS_HEADROOM_CEILING or net_mag > NET_HEADROOM_CEILING
        if breached:
            print(
                f"  [risk] projected gross={gross:,.0f} (>{GROSS_HEADROOM_CEILING}) "
                f"|net_sum|={net_mag:,.0f} (>{NET_HEADROOM_CEILING})",
                flush=True,
            )
        return breached

    # ---------- Session time ----------

    @staticmethod
    def _ticks_remaining_session(case: dict[str, Any]) -> int:
        tpp = int(case.get("ticks_per_period") or 300)
        tp = int(case.get("total_periods") or 1)
        period = int(case.get("period") or 1)
        tick = int(case.get("tick") or 0)
        total_ticks = tp * tpp
        elapsed = (period - 1) * tpp + tick
        return max(0, total_ticks - elapsed)

    # ---------- Unwind ----------

    def _max_chunk(self, ticker: str) -> int:
        return MAX_ORDER_QTY.get(ticker, DEFAULT_MAX_QTY)

    def _allowed_liability_qty(
        self,
        ticker: str,
        action: ActionEnum,
        unwind: UnwindState | None = None,
    ) -> int:
        """
        Strict liability-reduction guard using accepted tender liability first,
        with position-based cap when a consistent position signal is available.
        """
        if unwind is not None:
            if action != unwind.unwind_action:
                return 0
            liability_allowed = max(0, int(unwind.liability_remaining))
            if liability_allowed <= 0:
                return 0
            pos = self._position_qty(ticker)
            if action == ActionEnum.SELL:
                pos_cap = max(0, pos)
            else:
                pos_cap = max(0, -pos)
            # Position snapshots can lag immediately after accept_tender; only
            # cap by position when signal is available.
            if pos_cap > 0:
                return min(liability_allowed, pos_cap)
            return liability_allowed

        pos = self._position_qty(ticker)
        if action == ActionEnum.SELL:
            return max(0, pos)
        if action == ActionEnum.BUY:
            return max(0, -pos)
        return 0

    def _reconcile_liability_from_position(self, unwind: UnwindState) -> None:
        """Reduce remaining liability when observed position indicates progress."""
        pos = self._position_qty(unwind.ticker)
        if unwind.unwind_action == ActionEnum.BUY:
            if pos >= 0:
                return
            observed_remaining = max(0, -pos)
        else:
            if pos <= 0:
                return
            observed_remaining = max(0, pos)
        unwind.liability_remaining = min(unwind.liability_remaining, observed_remaining)

    def _market_chunk_orders(
        self,
        ticker: str,
        action: ActionEnum,
        total_qty: int,
        unwind: UnwindState | None = None,
        label: str = "unwind",
    ) -> None:
        max_q = self._max_chunk(ticker)
        remaining = total_qty
        while remaining > 0:
            allowed = self._allowed_liability_qty(ticker, action, unwind=unwind)
            if allowed <= 0:
                print(
                    f"    [liability-block] {label} {action.value} {ticker}: no liability-backed qty available",
                    flush=True,
                )
                break
            chunk = min(remaining, max_q, allowed)
            if chunk <= 0:
                break
            try:
                order = self.client.place_order(ticker, OrderType.MARKET, chunk, action)
                fill_price = float(
                    order.get("price")
                    or order.get("fill_price")
                    or order.get("avg_price")
                    or 0
                )
                if fill_price <= 0:
                    fill_price = self._get_last_trade_price(ticker)
                    if fill_price <= 0:
                        fill_price = self._get_mid(ticker) or 0.0
                mid_now = self._get_mid(ticker) or 0.0
                if mid_now > 0 and fill_price > 0 and abs(fill_price - mid_now) / mid_now > 0.05:
                    print(
                        f"    [warn] suspicious fill for {ticker}: fill={fill_price:.4f} mid={mid_now:.4f}",
                        flush=True,
                    )
                left_after = remaining - chunk
                print(
                    f"    [{label}] {action.value} {chunk} {ticker} @ {fill_price:.4f} "
                    f"({left_after} left)",
                    flush=True,
                )
                if unwind is not None:
                    unwind.unwind_value_accum += fill_price * chunk
                    unwind.unwind_qty_accum += chunk
                    unwind.liability_remaining = max(0, unwind.liability_remaining - chunk)
            except RITError as e:
                print(f"    [error] market unwind failed: {e}", flush=True)
                break
            remaining -= chunk
            time.sleep(0.1)

    def _place_limit_slices(
        self,
        ticker: str,
        action: ActionEnum,
        qty: float,
        price: float,
        unwind: UnwindState | None = None,
    ) -> list[int]:
        ids: list[int] = []
        remaining = int(qty)
        max_q = self._max_chunk(ticker)
        while remaining > 0:
            allowed = self._allowed_liability_qty(ticker, action, unwind=unwind)
            if allowed <= 0:
                print(
                    f"    [liability-block] unwind-limit {action.value} {ticker}: no liability-backed qty available",
                    flush=True,
                )
                break
            chunk = min(remaining, max_q, allowed)
            if chunk <= 0:
                break
            try:
                od = self.client.place_order(ticker, OrderType.LIMIT, chunk, action, price=price)
                oid = int(od.get("order_id") or od.get("id") or 0)
                if oid:
                    ids.append(oid)
                print(f"    [unwind-limit] order {oid} {action.value} {chunk} @ {price:.4f}", flush=True)
            except RITError as e:
                print(f"    [error] limit unwind failed @ {price}: {e}", flush=True)
                break
            remaining -= chunk
        return ids

    def _cancel_limits(self, order_ids: list[int]) -> None:
        for oid in order_ids:
            try:
                self.client.cancel_order(oid)
            except RITError as e:
                print(f"    [warn] cancel_order({oid}) failed: {e}", flush=True)

    def _start_unwind(
        self,
        tender_id: int,
        ticker: str,
        quantity: int,
        tender_price: float,
        tender_was_buy: bool,
    ) -> UnwindState:
        unwind_action = ActionEnum.SELL if tender_was_buy else ActionEnum.BUY
        mr = MARKET_ORDER_RATIO
        mq = min(quantity, max(1, round(quantity * mr)))
        remain = quantity - mq

        print(
            f"  [ACCEPTED] tender {tender_id} — unwind {quantity:,} {ticker} "
            f"(market tranche ~{mq}, limit tranche ~{remain})",
            flush=True,
        )

        case = self.client.get_case()
        u = UnwindState(
            tender_id=tender_id,
            ticker=ticker,
            quantity=quantity,
            tender_price=tender_price,
            tender_was_buy=tender_was_buy,
            unwind_action=unwind_action,
            limit_order_ids=[],
            last_mgmt_tick=int(case.get("tick", 0) or 0),
            liability_remaining=quantity,
        )

        print(
            f"  [unwind-debug] tender {tender_id} start pos={self._position_qty(ticker):,} "
            f"liab_rem={u.liability_remaining:,} action={unwind_action.value}",
            flush=True,
        )
        self._market_chunk_orders(ticker, unwind_action, mq, unwind=u)

        bids, asks = self._best_bb(ticker)
        tick_sz = self._infer_tick(bids, asks)

        # Strict liability-safe limit posting: only post up to currently backed remainder.
        self._reconcile_liability_from_position(u)
        remain = min(remain, self._allowed_liability_qty(ticker, unwind_action, unwind=u))

        if remain > 0:
            u.used_limit_orders = True
            if unwind_action == ActionEnum.SELL:
                if asks > 0:
                    limit_px = self._limit_sell_inside_spread(bids, asks, tick_sz)
                else:
                    m = self._get_mid(ticker) or bids
                    limit_px = max(tick_sz, round(float(m) - tick_sz, 4))
                u.limit_order_ids.extend(
                    self._place_limit_slices(ticker, ActionEnum.SELL, remain, limit_px, unwind=u)
                )
            else:
                if bids > 0:
                    limit_px = self._limit_buy_inside_spread(bids, asks, tick_sz)
                else:
                    m = self._get_mid(ticker) or asks
                    limit_px = round(float(m) + tick_sz, 4)
                u.limit_order_ids.extend(
                    self._place_limit_slices(ticker, ActionEnum.BUY, remain, limit_px, unwind=u)
                )

        return u

    def _open_limit_ids(self, ticker: str) -> list[int]:
        try:
            open_os = self.client.get_orders(status="OPEN")
        except Exception as e:
            print(f"  [warn] get_orders OPEN: {e}", flush=True)
            return []
        out: list[int] = []
        for o in open_os:
            if str(o.get("ticker")) == ticker and str(o.get("type") or "").upper() == "LIMIT":
                oid = o.get("order_id") or o.get("id")
                if oid is not None:
                    out.append(int(oid))
        return out

    def _position_qty(self, ticker: str) -> int:
        try:
            pos = self.client.get_positions().get(ticker)
            return int(pos.position) if pos else 0
        except Exception:
            return 0

    def _flatten_market_all(self, ticker: str, action: ActionEnum, unwind: UnwindState | None) -> None:
        q = self._allowed_liability_qty(ticker, action, unwind=unwind)
        if q <= 0:
            return
        self._market_chunk_orders(ticker, action, q, unwind=unwind, label="flatten")

    def _pulse_unwind(self, case: dict[str, Any]) -> None:
        if self._unwind is None or self._unwind.done:
            return

        u = self._unwind
        tick = int(case.get("tick", 0) or 0)
        t_rem = self._ticks_remaining_session(case)
        self._reconcile_liability_from_position(u)

        pos = self._position_qty(u.ticker)
        open_limit_count = len(self._open_limit_ids(u.ticker))

        if u.liability_remaining <= 0 and pos == 0 and open_limit_count == 0:
            u.done = True
            print(f"  [unwind] tender {u.tender_id} position flat.", flush=True)
            return

        if u.liability_remaining <= 0 and pos == 0 and open_limit_count > 0:
            self._cancel_limits(u.limit_order_ids)
            u.limit_order_ids.clear()
            u.done = True
            print(f"  [unwind] tender {u.tender_id} position flat (cancelled resting limits).", flush=True)
            return

        if tick - u.last_mgmt_tick < 10:
            return

        u.last_mgmt_tick = tick

        if t_rem < TIME_DECAY_THRESHOLD:
            print(
                f"  [urgency] t_rem={t_rem} < {TIME_DECAY_THRESHOLD} — flatten {u.ticker}",
                flush=True,
            )
            self._cancel_limits(u.limit_order_ids)
            u.limit_order_ids.clear()
            self._flatten_market_all(u.ticker, u.unwind_action, u)
            u.done = True
            return

        bids, asks = self._best_bb(u.ticker)
        tick_sz = self._infer_tick(bids, asks)
        self._cancel_limits(u.limit_order_ids)
        u.limit_order_ids.clear()

        rem = self._allowed_liability_qty(u.ticker, u.unwind_action, unwind=u)
        if rem <= 0:
            print(
                f"  [unwind-wait] tender {u.tender_id} awaiting liability visibility "
                f"(liab_rem={u.liability_remaining:,}, pos={pos:,})",
                flush=True,
            )
            return

        if u.unwind_action == ActionEnum.SELL:
            limit_px = (
                self._limit_sell_inside_spread(bids, asks, tick_sz)
                if asks > 0
                else max(tick_sz, round((self._get_mid(u.ticker) or bids) - tick_sz, 4))
            )
            u.limit_order_ids.extend(
                self._place_limit_slices(u.ticker, ActionEnum.SELL, rem, float(limit_px), unwind=u)
            )
        else:
            limit_px = (
                self._limit_buy_inside_spread(bids, asks, tick_sz)
                if bids > 0
                else round((self._get_mid(u.ticker) or 0) + tick_sz, 4)
            )
            u.limit_order_ids.extend(
                self._place_limit_slices(u.ticker, ActionEnum.BUY, rem, float(limit_px), unwind=u)
            )

    def _apply_pnl_record(self, rec: TenderRecord, u: UnwindState) -> None:
        """Close any residual inventory and populate realized P&L."""

        ticker = rec.ticker
        rq = abs(self._position_qty(ticker))
        if rq > 0:
            print(f"  [warn] residual position {rq} — market flatten.", flush=True)
            self._flatten_market_all(ticker, u.unwind_action, u)

        qty = rec.quantity
        filled_qty = max(0, min(qty, u.unwind_qty_accum))
        total_val = u.unwind_value_accum
        denom = filled_qty if filled_qty else 1
        avg_fill = total_val / denom
        residual = abs(self._position_qty(ticker))
        filled_eff = filled_qty

        if rec.action and "BUY" in rec.action.upper():
            pnl_gross = (avg_fill - float(rec.tender_price)) * max(0, filled_eff)
        else:
            pnl_gross = (float(rec.tender_price) - avg_fill) * max(0, filled_eff)

        unwind_filled_qty = u.unwind_qty_accum
        liability_violation_volume = max(0, int(unwind_filled_qty - qty))
        commission_cost = unwind_filled_qty * COMMISSION_PER_SHARE
        liability_fine = liability_violation_volume * LIABILITY_FINE_PER_SHARE
        pnl_net = pnl_gross - commission_cost - liability_fine

        rec.fill_price = avg_fill
        rec.pnl = round(pnl_gross, 4)
        rec.commission_cost = round(commission_cost, 4)
        rec.liability_violation_volume = liability_violation_volume
        rec.liability_fine = round(liability_fine, 4)
        rec.net_pnl = round(pnl_net, 4)
        rec.residual = residual
        print(
            f"  [P&L] tender {rec.tender_id}: filled~={filled_eff:,} residual={residual:,} "
            f"avg_unwind={avg_fill:.4f} gross={pnl_gross:+.2f} "
            f"comm={commission_cost:.2f} liab_vol={liability_violation_volume:,} "
            f"liab_fine={liability_fine:.2f} net={pnl_net:+.2f}",
            flush=True,
        )

    def _run_unwind_to_completion(self, rec: TenderRecord) -> None:
        """Pump the unwind state machine until flat or urgency/timeout."""

        deadline = time.monotonic() + 900.0
        while self._unwind is not None and not self._unwind.done:
            case = self.client.get_case()
            if case.get("status") == "STOPPED":
                break
            if time.monotonic() > deadline:
                print(f"  [warn] unwind timeout — forcing urgency flatten.", flush=True)
                u = self._unwind
                self._cancel_limits(u.limit_order_ids)
                u.limit_order_ids.clear()
                self._flatten_market_all(u.ticker, u.unwind_action, u)
                u.done = True
                break

            self._pulse_unwind(case)
            time.sleep(POLL_INTERVAL_SEC)

        if self._unwind is not None:
            u = self._unwind
            if not u.done:
                self._cancel_limits(u.limit_order_ids)
                u.limit_order_ids.clear()
                self._flatten_market_all(u.ticker, u.unwind_action, u)
                u.done = True
            self._apply_pnl_record(rec, u)
            rec.unwind_method = "limit" if u.used_limit_orders else "market"
            self._unwind = None

    # ---------- Tender evaluation ----------

    @staticmethod
    def _parse_tender_input(tender: Any) -> TenderInput | None:
        tender_id = getattr(tender, "id", None)
        ticker = getattr(tender, "ticker", None)
        quantity = getattr(tender, "quantity", 0)
        tender_price = float(getattr(tender, "price", 0) or 0)
        expiry_tick = int(getattr(tender, "expiry_tick", 0) or 0)
        action_raw = getattr(tender, "action", None)
        if not all([tender_id, ticker, quantity, tender_price, action_raw]):
            return None
        action_str = str(action_raw).upper()
        return TenderInput(
            tender_id=int(tender_id),
            ticker=str(ticker),
            quantity=int(quantity),
            tender_price=tender_price,
            expiry_tick=expiry_tick,
            action_str=action_str,
            tender_is_buy="BUY" in action_str,
        )

    def _decline_for_no_trade_window(self, tender_input: TenderInput, tick: int, time_remaining: int) -> TenderRecord:
        print(
            f"  [skip] tender {tender_input.tender_id}: time_remaining={time_remaining} "
            f"<= {NO_TRADE_TICKS_BEFORE_EXPIRY} (no-trade window)",
            flush=True,
        )
        try:
            self.client.decline_tender(tender_input.tender_id)
        except RITError as e:
            print(f"  [warn] decline_tender({tender_input.tender_id}) failed: {e}", flush=True)
        return TenderRecord(
            tick=tick,
            tender_id=tender_input.tender_id,
            ticker=tender_input.ticker,
            action=tender_input.action_str,
            quantity=tender_input.quantity,
            tender_price=tender_input.tender_price,
            mid_price=0.0,
            edge=0.0,
            decision="DECLINED",
            unwind_method="",
            risk_blocked=False,
        )

    def _load_market_snapshot(self, tender_input: TenderInput) -> TenderMarketSnapshot | None:
        mid = self._get_mid(tender_input.ticker)
        if mid is None or mid == 0:
            print(
                f"  [skip] tender {tender_input.tender_id}: could not get mid for {tender_input.ticker}",
                flush=True,
            )
            return None

        try:
            book = self.client.get_securities_book(tender_input.ticker, limit=BOOK_DEPTH)
        except RITError as e:
            print(f"  [skip] tender {tender_input.tender_id}: book error {e}", flush=True)
            return None

        bids = sorted(book.bids, key=lambda x: float(x.price), reverse=True)
        asks = sorted(book.asks, key=lambda x: float(x.price))
        bid_depth_total = sum(self._level_avail(l) for l in bids[:BOOK_DEPTH])
        ask_depth_total = sum(self._level_avail(l) for l in asks[:BOOK_DEPTH])
        n_bid_levels = sum(1 for l in bids[:BOOK_DEPTH] if self._level_avail(l) > 0)
        best_bid = float(bids[0].price) if bids else 0.0
        best_ask = float(asks[0].price) if asks else 0.0

        if tender_input.tender_is_buy:
            if not bids:
                print(f"  [skip] tender {tender_input.tender_id}: empty bid book", flush=True)
                return None
            best_bid = float(bids[0].price)
            slip, ok = self._estimate_slippage_buy_tender(bids, tender_input.quantity, best_bid)
        else:
            if not asks:
                print(f"  [skip] tender {tender_input.tender_id}: empty ask book", flush=True)
                return None
            best_ask = float(asks[0].price)
            slip, ok = self._estimate_slippage_sell_tender(asks, tender_input.quantity, best_ask)

        unwind_side_depth = bid_depth_total if tender_input.tender_is_buy else ask_depth_total
        if not ok or not math.isfinite(slip):
            slip = float("inf")
        return TenderMarketSnapshot(
            mid=mid,
            bid_depth_total=bid_depth_total,
            ask_depth_total=ask_depth_total,
            n_bid_levels=n_bid_levels,
            best_bid=best_bid,
            best_ask=best_ask,
            slip=slip,
            book_covers_full_qty=unwind_side_depth >= tender_input.quantity,
        )

    @staticmethod
    def _compute_tender_decision(
        tender_input: TenderInput,
        snapshot: TenderMarketSnapshot,
        risk_blocked: bool,
    ) -> TenderDecision:
        edge = (snapshot.mid - tender_input.tender_price) / snapshot.mid
        spread_captured = abs(snapshot.mid - tender_input.tender_price) * tender_input.quantity
        txn_cost = tender_input.quantity * COMMISSION_PER_SHARE
        expected_pnl = spread_captured - snapshot.slip - txn_cost
        liability_safe = snapshot.book_covers_full_qty and math.isfinite(snapshot.slip)
        decline_reason = None
        if risk_blocked:
            decline_reason = "risk limit"
        elif not liability_safe:
            decline_reason = "liability safety gate"
        elif expected_pnl <= 0:
            decline_reason = "E[pnl] <= 0"
        return TenderDecision(
            edge=edge,
            spread_captured=spread_captured,
            txn_cost=txn_cost,
            expected_pnl=expected_pnl,
            risk_blocked=risk_blocked,
            liability_safe=liability_safe,
            decline_reason=decline_reason,
        )

    def _log_tender_diagnostics(
        self,
        tender_input: TenderInput,
        snapshot: TenderMarketSnapshot,
        decision: TenderDecision,
    ) -> None:
        print(
            f"  [tender {tender_input.tender_id}] {tender_input.ticker} | action={tender_input.action_str} "
            f"| qty={tender_input.quantity:,} | tender_price={tender_input.tender_price:.4f} "
            f"| mid={snapshot.mid:.4f} | edge={decision.edge:.4f} | slip~={snapshot.slip:.2f} "
            f"| spread_cap~={decision.spread_captured:.2f} | txn={decision.txn_cost:.2f} "
            f"| E[pnl]~={decision.expected_pnl:.2f} | liab_safe={decision.liability_safe} "
            f"| risk_block={decision.risk_blocked}",
            flush=True,
        )

    @staticmethod
    def _slippage_for_record(slip: float) -> float:
        return slip if math.isfinite(slip) else -1.0

    def _decline_tender_with_record(
        self,
        tick: int,
        tender_input: TenderInput,
        snapshot: TenderMarketSnapshot,
        decision: TenderDecision,
    ) -> TenderRecord:
        try:
            self.client.decline_tender(tender_input.tender_id)
        except RITError as e:
            print(f"  [warn] decline_tender({tender_input.tender_id}) failed: {e}", flush=True)
        print(f"  [DECLINED] tender {tender_input.tender_id} — {decision.decline_reason}", flush=True)
        slip_value = self._slippage_for_record(snapshot.slip)
        expected_value = decision.expected_pnl if math.isfinite(decision.expected_pnl) else -1.0
        return TenderRecord(
            tick=tick,
            tender_id=tender_input.tender_id,
            ticker=tender_input.ticker,
            action=tender_input.action_str,
            quantity=tender_input.quantity,
            tender_price=tender_input.tender_price,
            mid_price=snapshot.mid,
            edge=round(decision.edge, 6),
            decision="DECLINED",
            bid_depth_total=snapshot.bid_depth_total,
            ask_depth_total=snapshot.ask_depth_total,
            n_bid_levels=snapshot.n_bid_levels,
            best_bid=snapshot.best_bid,
            best_ask=snapshot.best_ask,
            estimated_slippage=slip_value,
            book_covers_full_qty=snapshot.book_covers_full_qty,
            unwind_method="",
            spread_captured=decision.spread_captured,
            slippage_est=slip_value,
            txn_cost=decision.txn_cost,
            expected_pnl=expected_value,
            risk_blocked=decision.risk_blocked,
        )

    def _accept_tender_and_run_unwind(
        self,
        tick: int,
        tender_input: TenderInput,
        snapshot: TenderMarketSnapshot,
        decision: TenderDecision,
    ) -> TenderRecord | None:
        try:
            self.client.accept_tender(tender_input.tender_id)
        except RITError as e:
            print(f"  [error] accept_tender({tender_input.tender_id}) failed: {e}", flush=True)
            return None

        slip_value = self._slippage_for_record(snapshot.slip)
        rec = TenderRecord(
            tick=tick,
            tender_id=tender_input.tender_id,
            ticker=tender_input.ticker,
            action=tender_input.action_str,
            quantity=tender_input.quantity,
            tender_price=tender_input.tender_price,
            mid_price=snapshot.mid,
            edge=round(decision.edge, 6),
            decision="ACCEPTED",
            bid_depth_total=snapshot.bid_depth_total,
            ask_depth_total=snapshot.ask_depth_total,
            n_bid_levels=snapshot.n_bid_levels,
            best_bid=snapshot.best_bid,
            best_ask=snapshot.best_ask,
            estimated_slippage=slip_value,
            book_covers_full_qty=snapshot.book_covers_full_qty,
            spread_captured=decision.spread_captured,
            slippage_est=slip_value,
            txn_cost=decision.txn_cost,
            expected_pnl=decision.expected_pnl,
            risk_blocked=False,
        )
        self._unwind = self._start_unwind(
            tender_input.tender_id,
            tender_input.ticker,
            tender_input.quantity,
            tender_input.tender_price,
            tender_input.tender_is_buy,
        )
        self._run_unwind_to_completion(rec)
        return rec

    def evaluate_tender(self, tender: Any, tick: int) -> TenderRecord | None:
        tender_input = self._parse_tender_input(tender)
        if tender_input is None:
            return None

        time_remaining = tender_input.expiry_tick - tick
        if time_remaining <= NO_TRADE_TICKS_BEFORE_EXPIRY:
            return self._decline_for_no_trade_window(tender_input, tick, time_remaining)

        snapshot = self._load_market_snapshot(tender_input)
        if snapshot is None:
            return None

        risk_blocked = self._projected_exposure_breach(
            tender_input.ticker,
            tender_input.quantity,
            tender_input.tender_is_buy,
        )
        decision = self._compute_tender_decision(
            tender_input=tender_input,
            snapshot=snapshot,
            risk_blocked=risk_blocked,
        )
        self._log_tender_diagnostics(tender_input, snapshot, decision)

        if decision.decline_reason is not None:
            return self._decline_tender_with_record(
                tick=tick,
                tender_input=tender_input,
                snapshot=snapshot,
                decision=decision,
            )
        return self._accept_tender_and_run_unwind(
            tick=tick,
            tender_input=tender_input,
            snapshot=snapshot,
            decision=decision,
        )

    # ---------- Session loop ----------

    def run_session(self, session_num: int) -> SessionStats:
        stats = SessionStats(session=session_num)
        seen_tenders: set[int] = set()

        print(f"\n[session {session_num}] Starting tick loop...", flush=True)

        while True:
            case = self.client.get_case()
            status = case.get("status", "")
            tick = int(case.get("tick", 0) or 0)

            if status == "STOPPED":
                print(f"[session {session_num}] STOPPED — session complete.", flush=True)
                break

            self._pulse_unwind(case)

            if self._unwind is None:
                try:
                    tenders_raw = self.client.get_tenders()
                except Exception as e:
                    print(f"  [warn] get_tenders failed: {e}", flush=True)
                    tenders_raw = None

                if tenders_raw:
                    for tender in tenders_raw:
                        tid = getattr(tender, "id", None)
                        if tid is None or tid in seen_tenders:
                            continue
                        seen_tenders.add(tid)
                        stats.tenders_seen += 1

                        record = self.evaluate_tender(tender, tick)
                        if record:
                            stats.records.append(record)
                            if record.decision == "ACCEPTED":
                                stats.tenders_accepted += 1
                                stats.total_gross_pnl += record.pnl
                                stats.total_pnl += record.net_pnl
                            else:
                                stats.tenders_declined += 1

            if tick >= 299:
                if self._unwind is not None:
                    print(f"  [warn] End of ticks — forcing unwind flatten.", flush=True)
                    u = self._unwind
                    self._cancel_limits(u.limit_order_ids)
                    u.limit_order_ids.clear()
                    self._flatten_market_all(u.ticker, u.unwind_action, u)
                    u.done = True
                    self._unwind = None
                print(f"[session {session_num}] Final tick reached.", flush=True)
                break

            time.sleep(POLL_INTERVAL_SEC)

        return stats


def _wait_for_session_start(client: RotmanSDK) -> bool:
    return wait_for_session_start(client=client, stopped_wait_timeout=STOPPED_WAIT_TIMEOUT)


def _write_session_xlsx(all_stats: list[SessionStats], out_path: Path) -> None:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = [
        "Session",
        "Tenders Seen",
        "Accepted",
        "Declined",
        "Accept Rate",
        "Total Gross P&L",
        "Total Net P&L",
    ]
    ws_sum.append(sum_headers)
    for s in all_stats:
        rate = f"{s.tenders_accepted / s.tenders_seen:.1%}" if s.tenders_seen else "0.0%"
        ws_sum.append(
            [
                s.session,
                s.tenders_seen,
                s.tenders_accepted,
                s.tenders_declined,
                rate,
                round(s.total_gross_pnl, 4),
                round(s.total_pnl, 4),
            ]
        )
    n = len(all_stats)
    if n > 1:
        total_gross_pnl = sum(ws.total_gross_pnl for ws in all_stats)
        total_net_pnl = sum(ws.total_pnl for ws in all_stats)
        total_seen = sum(ws.tenders_seen for ws in all_stats)
        total_acc = sum(ws.tenders_accepted for ws in all_stats)
        total_decl = sum(ws.tenders_declined for ws in all_stats)
        overall = f"{total_acc / total_seen:.1%}" if total_seen else "0.0%"
        ws_sum.append(["TOTAL", total_seen, total_acc, total_decl, overall, round(total_gross_pnl, 4), round(total_net_pnl, 4)])
        for cell in ws_sum[ws_sum.max_row]:
            cell.font = Font(name="Arial", bold=True)
    _style_ws(ws_sum, len(sum_headers))

    ws_rec = wb.create_sheet("Tender Detail")
    rec_headers = [
        "Session",
        "Tick",
        "Tender ID",
        "Ticker",
        "Action",
        "Quantity",
        "Tender Price",
        "Mid Price",
        "Edge",
        "Bid Depth Total",
        "Ask Depth Total",
        "N Bid Levels",
        "Best Bid",
        "Best Ask",
        "Estimated Slippage",
        "Book Covers Full Qty",
        "Spread Captured",
        "Slippage Est",
        "Txn Cost",
        "Expected PnL",
        "Risk blocked",
        "Decision",
        "Unwind Method",
        "Avg Unwind Price",
        "Gross P&L",
        "Commission Cost",
        "Liability Violation Volume",
        "Liability Fine",
        "Net P&L",
        "Residual",
    ]
    ws_rec.append(rec_headers)
    for st in all_stats:
        for r in st.records:
            ws_rec.append(
                [
                    st.session,
                    r.tick,
                    r.tender_id,
                    r.ticker,
                    r.action,
                    r.quantity,
                    r.tender_price,
                    r.mid_price,
                    r.edge,
                    r.bid_depth_total,
                    r.ask_depth_total,
                    r.n_bid_levels,
                    r.best_bid,
                    r.best_ask,
                    r.estimated_slippage,
                    r.book_covers_full_qty,
                    r.spread_captured,
                    r.slippage_est,
                    r.txn_cost,
                    r.expected_pnl,
                    r.risk_blocked,
                    r.decision,
                    r.unwind_method or "",
                    r.fill_price or "",
                    r.pnl or "",
                    r.commission_cost or "",
                    r.liability_violation_volume or "",
                    r.liability_fine or "",
                    r.net_pnl or "",
                    r.residual or "",
                ]
            )
    _style_ws(ws_rec, len(rec_headers))
    ws_rec.freeze_panes = "A2"

    green_fill = PatternFill("solid", start_color="C6EFCE")
    red_fill = PatternFill("solid", start_color="FFC7CE")
    decision_col = rec_headers.index("Decision") + 1
    for row in ws_rec.iter_rows(min_row=2, min_col=decision_col, max_col=decision_col):
        for cell in row:
            if cell.value == "ACCEPTED":
                cell.fill = green_fill
            elif cell.value == "DECLINED":
                cell.fill = red_fill

    wb.save(out_path)
    wb.close()


def main() -> None:
    client = RotmanSDK(API_KEY=API_KEY, HOST=HOST)
    agent = DeliberativeAgent(client)
    all_stats: list[SessionStats] = []
    session = 0

    try:
        while MAX_SESSIONS is None or session < MAX_SESSIONS:
            session += 1
            print(f"\n{'='*60}", flush=True)
            print(
                f"  SESSION {session}{f'/{MAX_SESSIONS}' if MAX_SESSIONS else ''}  | "
                f"mkt_ratio={MARKET_ORDER_RATIO} urgency<{TIME_DECAY_THRESHOLD} ticks",
                flush=True,
            )
            print(f"{'='*60}", flush=True)

            if not _wait_for_session_start(client):
                print(f"Timed out waiting for session {session}. Stopping.")
                break

            stats = agent.run_session(session)
            all_stats.append(stats)

            print(f"\n[session {session} summary]", flush=True)
            print(f"  Tenders seen:     {stats.tenders_seen}", flush=True)
            print(f"  Accepted:         {stats.tenders_accepted}", flush=True)
            print(f"  Declined:         {stats.tenders_declined}", flush=True)
            print(f"  Total Gross P&L:  {stats.total_gross_pnl:+.4f}", flush=True)
            print(f"  Total Net P&L:    {stats.total_pnl:+.4f}", flush=True)

    except KeyboardInterrupt:
        print("\nInterrupted — saving logs.", flush=True)

    if not all_stats:
        print("No sessions completed.")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"deliberative_agent_{run_id}.xlsx"
    _write_session_xlsx(all_stats, out_path)
    print(f"\nSaved: {out_path}", flush=True)


if __name__ == "__main__":
    main()
