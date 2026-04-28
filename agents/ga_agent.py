"""
GA-driven Deliberative Agent
----------------------------
Same behavior as deliberative_agent, but runtime parameters are loaded from
ga/evolved_params.json.
"""

from __future__ import annotations

import json
import math
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import deliberative_agent as base
from sdk import RITError, RotmanSDK


EVOLVED = json.loads(
    (Path(__file__).parent.parent / "ga" / "evolved_params.json").read_text()
)
PARAMS = EVOLVED["params"]

MARKET_ORDER_RATIO = PARAMS["market_order_ratio"]
URGENCY_THRESHOLD = int(300 * PARAMS["time_decay_factor"] / 3.0)
SLICE_SIZE_RATIO = PARAMS["slice_size_ratio"]

# Apply GA params to shared deliberative logic.
base.MARKET_ORDER_RATIO = MARKET_ORDER_RATIO
base.TIME_DECAY_THRESHOLD = URGENCY_THRESHOLD

API_KEY = base.API_KEY
HOST = base.HOST
MAX_SESSIONS = base.MAX_SESSIONS
OUTPUT_DIR = base.OUTPUT_DIR


class GAAgent(base.DeliberativeAgent):
    """DeliberativeAgent with GA-parametrized entry threshold and slice sizing."""

    def _place_limit_slices(
        self,
        ticker: str,
        action: base.ActionEnum,
        qty: float,
        price: float,
    ) -> list[int]:
        ids: list[int] = []
        remaining = int(qty)
        max_q = self._max_chunk(ticker)
        slice_q = max(1, int(max_q * float(SLICE_SIZE_RATIO)))
        while remaining > 0:
            chunk = min(remaining, slice_q)
            try:
                od = self.client.place_order(ticker, base.OrderType.LIMIT, chunk, action, price=price)
                oid = int(od.get("order_id") or od.get("id") or 0)
                if oid:
                    ids.append(oid)
                print(f"    [unwind-limit] order {oid} {action.value} {chunk} @ {price:.4f}", flush=True)
            except RITError as e:
                print(f"    [error] limit unwind failed @ {price}: {e}", flush=True)
                break
            remaining -= chunk
        return ids

    def evaluate_tender(self, tender: Any, tick: int) -> base.TenderRecord | None:
        tender_id = getattr(tender, "id", None)
        ticker = getattr(tender, "ticker", None)
        quantity = getattr(tender, "quantity", 0)
        tender_price = float(getattr(tender, "price", 0) or 0)
        expiry_tick = int(getattr(tender, "expiry_tick", 0) or 0)
        action_raw = getattr(tender, "action", None)

        if not all([tender_id, ticker, quantity, tender_price, action_raw]):
            return None

        action_str = str(action_raw).upper()
        tender_is_buy = "BUY" in action_str

        time_remaining = expiry_tick - tick
        if time_remaining <= base.NO_TRADE_TICKS_BEFORE_EXPIRY:
            print(
                f"  [skip] tender {tender_id}: time_remaining={time_remaining} "
                f"<= {base.NO_TRADE_TICKS_BEFORE_EXPIRY} (no-trade window)",
                flush=True,
            )
            try:
                self.client.decline_tender(tender_id)
            except RITError as e:
                print(f"  [warn] decline_tender({tender_id}) failed: {e}", flush=True)
            return base.TenderRecord(
                tick=tick,
                tender_id=tender_id,
                ticker=ticker,
                action=action_str,
                quantity=quantity,
                tender_price=tender_price,
                mid_price=0.0,
                edge=0.0,
                decision="DECLINED",
                unwind_method="",
                risk_blocked=False,
            )

        mid = self._get_mid(ticker)
        if mid is None or mid == 0:
            print(f"  [skip] tender {tender_id}: could not get mid for {ticker}", flush=True)
            return None

        edge = abs(mid - tender_price) / mid
        threshold = float(PARAMS["threshold"])

        try:
            book = self.client.get_securities_book(ticker, limit=base.BOOK_DEPTH)
        except RITError as e:
            print(f"  [skip] tender {tender_id}: book error {e}", flush=True)
            return None

        bids = sorted(book.bids, key=lambda x: float(x.price), reverse=True)
        asks = sorted(book.asks, key=lambda x: float(x.price))
        bid_depth_total = sum(self._level_avail(l) for l in bids[: base.BOOK_DEPTH])
        ask_depth_total = sum(self._level_avail(l) for l in asks[: base.BOOK_DEPTH])
        n_bid_levels = sum(1 for l in bids[: base.BOOK_DEPTH] if self._level_avail(l) > 0)
        best_bid = float(bids[0].price) if bids else 0.0
        best_ask = float(asks[0].price) if asks else 0.0

        if tender_is_buy:
            if not bids:
                print(f"  [skip] tender {tender_id}: empty bid book", flush=True)
                return None
            best_bid = float(bids[0].price)
            slip, ok = self._estimate_slippage_buy_tender(bids, quantity, best_bid)
        else:
            if not asks:
                print(f"  [skip] tender {tender_id}: empty ask book", flush=True)
                return None
            best_ask = float(asks[0].price)
            slip, ok = self._estimate_slippage_sell_tender(asks, quantity, best_ask)

        unwind_side_depth = bid_depth_total if tender_is_buy else ask_depth_total
        book_covers_full_qty = unwind_side_depth >= int(quantity)

        if not ok or not math.isfinite(slip):
            slip = float("inf")

        spread_captured = abs(mid - tender_price) * quantity
        txn_cost = quantity * 0.02
        expected_pnl = spread_captured - slip - txn_cost
        risk_blocked = self._projected_exposure_breach(ticker, quantity, tender_is_buy)
        accept = (edge > threshold) and (expected_pnl > 0) and (not risk_blocked)

        print(
            f"  [tender {tender_id}] {ticker} | action={action_str} | qty={quantity:,} | "
            f"tender_price={tender_price:.4f} | mid={mid:.4f} | edge={edge:.4f} "
            f"(thr={threshold:.4f}) | slip~={slip:.2f} | spread_cap~={spread_captured:.2f} | "
            f"txn={txn_cost:.2f} | E[pnl]~={expected_pnl:.2f} | risk_block={risk_blocked}",
            flush=True,
        )

        if not accept:
            try:
                self.client.decline_tender(tender_id)
            except RITError as e:
                print(f"  [warn] decline_tender({tender_id}) failed: {e}", flush=True)
            reason = "risk limit" if risk_blocked else "threshold/E[pnl] gate"
            print(f"  [DECLINED] tender {tender_id} — {reason}", flush=True)
            return base.TenderRecord(
                tick=tick,
                tender_id=tender_id,
                ticker=ticker,
                action=action_str,
                quantity=quantity,
                tender_price=tender_price,
                mid_price=mid,
                edge=round(edge, 6),
                decision="DECLINED",
                bid_depth_total=bid_depth_total,
                ask_depth_total=ask_depth_total,
                n_bid_levels=n_bid_levels,
                best_bid=best_bid,
                best_ask=best_ask,
                estimated_slippage=slip if math.isfinite(slip) else -1.0,
                book_covers_full_qty=book_covers_full_qty,
                unwind_method="",
                spread_captured=spread_captured,
                slippage_est=slip if math.isfinite(slip) else -1.0,
                txn_cost=txn_cost,
                expected_pnl=expected_pnl if math.isfinite(expected_pnl) else -1.0,
                risk_blocked=risk_blocked,
            )

        try:
            self.client.accept_tender(tender_id)
        except RITError as e:
            print(f"  [error] accept_tender({tender_id}) failed: {e}", flush=True)
            return None

        rec = base.TenderRecord(
            tick=tick,
            tender_id=tender_id,
            ticker=ticker,
            action=action_str,
            quantity=quantity,
            tender_price=tender_price,
            mid_price=mid,
            edge=round(edge, 6),
            decision="ACCEPTED",
            bid_depth_total=bid_depth_total,
            ask_depth_total=ask_depth_total,
            n_bid_levels=n_bid_levels,
            best_bid=best_bid,
            best_ask=best_ask,
            estimated_slippage=slip if math.isfinite(slip) else -1.0,
            book_covers_full_qty=book_covers_full_qty,
            spread_captured=spread_captured,
            slippage_est=slip if math.isfinite(slip) else -1.0,
            txn_cost=txn_cost,
            expected_pnl=expected_pnl,
            risk_blocked=False,
        )

        self._unwind = self._start_unwind(
            tender_id, ticker, quantity, tender_price, tender_is_buy
        )
        self._run_unwind_to_completion(rec)
        return rec


def _write_session_xlsx(all_stats: list[base.SessionStats], out_path: Path) -> None:
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = ["Session", "Tenders Seen", "Accepted", "Declined", "Accept Rate", "Total P&L", "Params"]
    ws_sum.append(sum_headers)
    params_str = json.dumps(PARAMS, sort_keys=True)
    for s in all_stats:
        rate = f"{s.tenders_accepted / s.tenders_seen:.1%}" if s.tenders_seen else "0.0%"
        ws_sum.append(
            [s.session, s.tenders_seen, s.tenders_accepted, s.tenders_declined, rate, round(s.total_pnl, 4), params_str]
        )
    if len(all_stats) > 1:
        total_pnl = sum(s.total_pnl for s in all_stats)
        total_seen = sum(s.tenders_seen for s in all_stats)
        total_acc = sum(s.tenders_accepted for s in all_stats)
        total_dec = sum(s.tenders_declined for s in all_stats)
        overall = f"{total_acc / total_seen:.1%}" if total_seen else "0.0%"
        ws_sum.append(["TOTAL", total_seen, total_acc, total_dec, overall, round(total_pnl, 4), params_str])
        for cell in ws_sum[ws_sum.max_row]:
            cell.font = Font(name="Arial", bold=True)
    base._style_ws(ws_sum, len(sum_headers))

    ws_rec = wb.create_sheet("Tender Detail")
    rec_headers = [
        "Session",
        "Tick",
        "Tender ID",
        "Ticker",
        "Action",
        "Quantity",
        "Tender Price",
        "Mid Price",
        "Edge",
        "Bid Depth Total",
        "Ask Depth Total",
        "N Bid Levels",
        "Best Bid",
        "Best Ask",
        "Estimated Slippage",
        "Book Covers Full Qty",
        "Spread Captured",
        "Slippage Est",
        "Txn Cost",
        "Expected PnL",
        "Risk blocked",
        "Decision",
        "Unwind Method",
        "Avg Unwind Price",
        "P&L",
        "Residual",
        "Params",
    ]
    ws_rec.append(rec_headers)
    for st in all_stats:
        for r in st.records:
            ws_rec.append(
                [
                    st.session,
                    r.tick,
                    r.tender_id,
                    r.ticker,
                    r.action,
                    r.quantity,
                    r.tender_price,
                    r.mid_price,
                    r.edge,
                    r.bid_depth_total,
                    r.ask_depth_total,
                    r.n_bid_levels,
                    r.best_bid,
                    r.best_ask,
                    r.estimated_slippage,
                    r.book_covers_full_qty,
                    r.spread_captured,
                    r.slippage_est,
                    r.txn_cost,
                    r.expected_pnl,
                    r.risk_blocked,
                    r.decision,
                    r.unwind_method or "",
                    r.fill_price or "",
                    r.pnl or "",
                    r.residual or "",
                    params_str,
                ]
            )
    base._style_ws(ws_rec, len(rec_headers))
    ws_rec.freeze_panes = "A2"

    green_fill = PatternFill("solid", start_color="C6EFCE")
    red_fill = PatternFill("solid", start_color="FFC7CE")
    decision_col = rec_headers.index("Decision") + 1
    for row in ws_rec.iter_rows(min_row=2, min_col=decision_col, max_col=decision_col):
        for cell in row:
            if cell.value == "ACCEPTED":
                cell.fill = green_fill
            elif cell.value == "DECLINED":
                cell.fill = red_fill

    wb.save(out_path)
    wb.close()


def main() -> None:
    print(f"[GA] loaded evolved params: {PARAMS}", flush=True)

    client = RotmanSDK(API_KEY=API_KEY, HOST=HOST)
    agent = GAAgent(client)
    all_stats: list[base.SessionStats] = []
    session = 0

    try:
        while MAX_SESSIONS is None or session < MAX_SESSIONS:
            session += 1
            print(f"\n{'='*60}", flush=True)
            print(
                f"  SESSION {session}{f'/{MAX_SESSIONS}' if MAX_SESSIONS else ''}  | "
                f"mkt_ratio={MARKET_ORDER_RATIO} urgency<{URGENCY_THRESHOLD} ticks "
                f"slice_ratio={SLICE_SIZE_RATIO}",
                flush=True,
            )
            print(f"{'='*60}", flush=True)

            if not base._wait_for_session_start(client):
                print(f"Timed out waiting for session {session}. Stopping.")
                break

            stats = agent.run_session(session)
            all_stats.append(stats)

            print(f"\n[session {session} summary]", flush=True)
            print(f"  Tenders seen:     {stats.tenders_seen}", flush=True)
            print(f"  Accepted:         {stats.tenders_accepted}", flush=True)
            print(f"  Declined:         {stats.tenders_declined}", flush=True)
            print(f"  Total P&L:        {stats.total_pnl:+.4f}", flush=True)

    except KeyboardInterrupt:
        print("\nInterrupted — saving logs.", flush=True)

    if not all_stats:
        print("No sessions completed.")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"ga_agent_{run_id}.xlsx"
    _write_session_xlsx(all_stats, out_path)
    print(f"\nSaved: {out_path}", flush=True)


if __name__ == "__main__":
    main()
