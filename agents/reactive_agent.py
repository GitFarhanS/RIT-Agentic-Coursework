"""
Phase 1 — Reactive Agent
------------------------
Monitors tenders each tick and accepts if:
    (mid_price - tender_price) / mid_price > THRESHOLD

On acceptance, immediately unwinds the resulting position using market orders,
respecting per-order maximums (CRZY: 25,000 / TAME: 10,000).

Logs per-session P&L, residual position, and tender accept/decline summary
to both stdout and an xlsx file in agent_logs/.
"""

from __future__ import annotations

import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from common_helpers import get_last_trade_price, get_mid_price, style_worksheet, wait_for_session_start
from runtime_config import load_runtime_config
from sdk import RITError, RotmanSDK
from utilities import ActionEnum, OrderType


# Config
_RUNTIME = load_runtime_config()
API_KEY = _RUNTIME.api_key
HOST = _RUNTIME.host

THRESHOLD = 0.01          # accept if (mid - tender_price) / mid > 2%
POLL_INTERVAL_SEC = 0.5   # poll faster than logger — tenders can be short-lived
MAX_SESSIONS = 10          # set to None to run forever
STOPPED_WAIT_TIMEOUT = 700

OUTPUT_DIR = Path(__file__).resolve().parent / "agent_logs"

MAX_ORDER_QTY: dict[str, int] = {
    "CRZY": 25_000,
    "TAME": 10_000,
}
DEFAULT_MAX_QTY = 10_000  # fallback for unknown tickers


# Data classes


@dataclass
class TenderRecord:
    tick: int
    tender_id: int
    ticker: str
    action: str       # BUY or SELL (what the tender asks US to do)
    quantity: int
    tender_price: float
    mid_price: float
    edge: float       # BUY: (mid - tender)/mid, SELL: (tender - mid)/mid
    decision: str     # "ACCEPTED" or "DECLINED"
    bid_depth_total: int = 0
    ask_depth_total: int = 0
    n_bid_levels: int = 0
    best_bid: float = 0.0
    best_ask: float = 0.0
    estimated_slippage: float = 0.0
    book_covers_full_qty: bool = False
    unwind_method: str = ""
    fill_price: float = 0.0    # actual unwind price (avg)
    pnl: float = 0.0           # realised P&L from this tender
    residual: int = 0          # shares left unhedged after unwind


@dataclass
class SessionStats:
    session: int
    tenders_seen: int = 0
    tenders_accepted: int = 0
    tenders_declined: int = 0
    total_pnl: float = 0.0
    records: list[TenderRecord] = field(default_factory=list)



# Styling helpers (shared with logger)


def _style_ws(ws, n_cols: int) -> None:
    style_worksheet(ws, n_cols=n_cols, max_col_width=35)



# ReactiveAgent


class ReactiveAgent:
    """
    Reactive tender agent for RIT LT3.

    Decision rule:
        BUY tender:  edge = (mid_price - tender_price) / mid_price
        SELL tender: edge = (tender_price - mid_price) / mid_price
        Accept if edge > threshold

    Unwind:
        After accepting a BUY tender  → sell the position with market orders
        After accepting a SELL tender → buy  the position with market orders
        Orders are chunked to respect per-ticker max order size.
    """

    def __init__(self, client: RotmanSDK, threshold: float = THRESHOLD) -> None:
        self.client = client
        self.threshold = threshold

    # ------------------------------------------------------------------
    # Mid-price
    # ------------------------------------------------------------------

    def _get_mid(self, ticker: str) -> float | None:
        """Return mid-price for ticker from current securities snapshot."""
        return get_mid_price(self.client, ticker)

    @staticmethod
    def _level_avail(level) -> int:
        return int(getattr(level, "quantity_remaining", 0) or getattr(level, "quantity", 0) or 0)

    def _get_last_trade_price(self, ticker: str) -> float:
        """Best-effort last traded price from securities endpoint."""
        return get_last_trade_price(self.client, ticker)

    def _snapshot_book_metrics(
        self, ticker: str, quantity: int, action_str: str
    ) -> tuple[int, int, int, float, float, float, bool]:
        """Return depth totals/top-of-book/slippage coverage snapshot at eval time."""
        try:
            book = self.client.get_securities_book(ticker, limit=20)
        except Exception as e:
            print(f"  [warn] book snapshot failed for {ticker}: {e}", flush=True)
            return 0, 0, 0, 0.0, 0.0, 0.0, False

        bids = sorted(book.bids, key=lambda x: float(x.price), reverse=True)
        asks = sorted(book.asks, key=lambda x: float(x.price))

        bid_depth_total = sum(self._level_avail(l) for l in bids[:20])
        ask_depth_total = sum(self._level_avail(l) for l in asks[:20])
        n_bid_levels = sum(1 for l in bids[:20] if self._level_avail(l) > 0)
        best_bid = float(bids[0].price) if bids else 0.0
        best_ask = float(asks[0].price) if asks else 0.0

        unwind_sell = "BUY" in action_str  # accepted BUY tender -> we SELL to unwind
        remaining = quantity
        estimated_slippage = 0.0
        if unwind_sell:
            for lvl in bids[:20]:
                px = float(lvl.price)
                av = self._level_avail(lvl)
                if av <= 0:
                    continue
                take = min(remaining, av)
                estimated_slippage += max(0.0, best_bid - px) * take
                remaining -= take
                if remaining <= 0:
                    break
        else:
            for lvl in asks[:20]:
                px = float(lvl.price)
                av = self._level_avail(lvl)
                if av <= 0:
                    continue
                take = min(remaining, av)
                estimated_slippage += max(0.0, px - best_ask) * take
                remaining -= take
                if remaining <= 0:
                    break

        book_covers_full_qty = remaining <= 0
        return (
            bid_depth_total,
            ask_depth_total,
            n_bid_levels,
            best_bid,
            best_ask,
            estimated_slippage,
            book_covers_full_qty,
        )

    # ------------------------------------------------------------------
    # Unwind
    # ------------------------------------------------------------------

    def _unwind(self, ticker: str, quantity: int, action: ActionEnum) -> tuple[float, int]:
        """
        Send market orders to unwind `quantity` shares in `action` direction.
        Chunks orders to respect per-ticker max size.
        Returns (avg_fill_price, residual_qty).
        """
        max_qty = MAX_ORDER_QTY.get(ticker, DEFAULT_MAX_QTY)
        remaining = quantity
        total_value = 0.0
        total_filled = 0

        while remaining > 0:
            chunk = min(remaining, max_qty)
            try:
                order = self.client.place_order(ticker, OrderType.MARKET, chunk, action)
                fill_price = float(order.get("price") or order.get("fill_price") or order.get("avg_price") or 0)
                if fill_price <= 0:
                    # RIT can return 0 for market orders. Retry from last trade, then mid.
                    fill_price = self._get_last_trade_price(ticker)
                    if fill_price <= 0:
                        fill_price = self._get_mid(ticker) or 0.0
                mid_now = self._get_mid(ticker) or 0.0
                if mid_now > 0 and fill_price > 0 and abs(fill_price - mid_now) / mid_now > 0.05:
                    print(
                        f"    [warn] suspicious fill for {ticker}: fill={fill_price:.4f} mid={mid_now:.4f}",
                        flush=True,
                    )
                total_value += fill_price * chunk
                total_filled += chunk
                remaining -= chunk
                print(
                    f"    [unwind] {action.value} {chunk} {ticker} @ {fill_price:.4f} "
                    f"({remaining} remaining)",
                    flush=True,
                )
            except RITError as e:
                print(f"    [error] unwind order failed: {e}", flush=True)
                break
            time.sleep(0.1)  # brief pause between chunks

        avg_price = total_value / total_filled if total_filled > 0 else 0.0
        return avg_price, remaining  # remaining = residual not filled

    # ------------------------------------------------------------------
    # Tender decision
    # ------------------------------------------------------------------

    def evaluate_tender(self, tender, tick: int) -> TenderRecord | None:
        """
        Evaluate a single tender. Returns a TenderRecord (accepted or declined),
        or None if tender data is incomplete.
        """
        tender_id = getattr(tender, "id", None) or getattr(tender, "tender_id", None)
        ticker = getattr(tender, "ticker", None)
        quantity = getattr(tender, "quantity", 0)
        tender_price = float(getattr(tender, "price", 0) or 0)
        action_raw = getattr(tender, "action", None)

        if not all([tender_id, ticker, quantity, tender_price, action_raw]):
            return None

        mid = self._get_mid(ticker)
        if mid is None or mid == 0:
            print(f"  [skip] tender {tender_id}: could not get mid for {ticker}", flush=True)
            return None

        action_str = str(action_raw).upper()
        if "BUY" in action_str:
            edge = (mid - tender_price) / mid
        else:
            edge = (tender_price - mid) / mid

        (
            bid_depth_total,
            ask_depth_total,
            n_bid_levels,
            best_bid,
            best_ask,
            estimated_slippage,
            book_covers_full_qty,
        ) = self._snapshot_book_metrics(ticker, int(quantity), action_str)

        print(
            f"  [tender {tender_id}] {ticker} | action={action_raw} | qty={quantity:,} | "
            f"tender_price={tender_price:.4f} | mid={mid:.4f} | edge={edge:.4f} "
            f"(threshold={self.threshold})",
            flush=True,
        )

        if edge > self.threshold:
            # Accept
            try:
                self.client.accept_tender(tender_id)
            except RITError as e:
                print(f"  [error] accept_tender({tender_id}) failed: {e}", flush=True)
                return None

            print(f"  [ACCEPTED] tender {tender_id} — unwinding {quantity:,} {ticker}", flush=True)

            # Unwind: if tender had us BUY → we now hold long → SELL to unwind
            # If tender had us SELL → we are short → BUY to unwind
            unwind_action = ActionEnum.SELL if "BUY" in action_str else ActionEnum.BUY
            avg_fill, residual = self._unwind(ticker, quantity, unwind_action)

            # P&L: we bought at tender_price and sold at avg_fill (or vice versa)
            filled_qty = quantity - residual
            if "BUY" in action_str:
                pnl = (avg_fill - tender_price) * filled_qty
            else:
                pnl = (tender_price - avg_fill) * filled_qty

            print(
                f"  [P&L] tender {tender_id}: filled={filled_qty:,} residual={residual:,} "
                f"avg_unwind={avg_fill:.4f} pnl={pnl:+.2f}",
                flush=True,
            )

            return TenderRecord(
                tick=tick, tender_id=tender_id, ticker=ticker,
                action=action_str, quantity=quantity,
                tender_price=tender_price, mid_price=mid, edge=round(edge, 6),
                decision="ACCEPTED",
                bid_depth_total=bid_depth_total,
                ask_depth_total=ask_depth_total,
                n_bid_levels=n_bid_levels,
                best_bid=best_bid,
                best_ask=best_ask,
                estimated_slippage=estimated_slippage,
                book_covers_full_qty=book_covers_full_qty,
                unwind_method="market",
                fill_price=avg_fill,
                pnl=round(pnl, 4), residual=residual,
            )
        else:
            # Decline
            try:
                self.client.decline_tender(tender_id)
            except RITError as e:
                print(f"  [warn] decline_tender({tender_id}) failed: {e}", flush=True)

            print(f"  [DECLINED] tender {tender_id} — edge {edge:.4f} < threshold {self.threshold}", flush=True)

            return TenderRecord(
                tick=tick, tender_id=tender_id, ticker=ticker,
                action=str(action_raw).upper(), quantity=quantity,
                tender_price=tender_price, mid_price=mid, edge=round(edge, 6),
                bid_depth_total=bid_depth_total,
                ask_depth_total=ask_depth_total,
                n_bid_levels=n_bid_levels,
                best_bid=best_bid,
                best_ask=best_ask,
                estimated_slippage=estimated_slippage,
                book_covers_full_qty=book_covers_full_qty,
                decision="DECLINED",
            )

    # ------------------------------------------------------------------
    # Session loop
    # ------------------------------------------------------------------

    def run_session(self, session_num: int) -> SessionStats:
        """Run one full session (tick 0 → 299). Returns SessionStats."""
        stats = SessionStats(session=session_num)
        seen_tenders: set[int] = set()

        print(f"\n[session {session_num}] Starting tick loop...", flush=True)

        while True:
            case = self.client.get_case()
            status = case.get("status", "")
            tick = int(case.get("tick", 0) or 0)

            if status == "STOPPED":
                print(f"[session {session_num}] STOPPED — session complete.", flush=True)
                break

            # Check for tenders
            try:
                tenders_raw = self.client.get_tenders()
            except Exception as e:
                print(f"  [warn] get_tenders failed: {e}", flush=True)
                tenders_raw = None

            if tenders_raw:
                for tender in tenders_raw:
                    tid = getattr(tender, "id", None) or getattr(tender, "tender_id", None)
                    if tid is None or tid in seen_tenders:
                        continue
                    seen_tenders.add(tid)
                    stats.tenders_seen += 1

                    record = self.evaluate_tender(tender, tick)
                    if record:
                        stats.records.append(record)
                        if record.decision == "ACCEPTED":
                            stats.tenders_accepted += 1
                            stats.total_pnl += record.pnl
                        else:
                            stats.tenders_declined += 1

            if tick >= 299:
                print(f"[session {session_num}] Final tick reached.", flush=True)
                break

            time.sleep(POLL_INTERVAL_SEC)

        return stats



# Session start detection (same proven logic as logger)


def _wait_for_session_start(client: RotmanSDK) -> bool:
    return wait_for_session_start(client=client, stopped_wait_timeout=STOPPED_WAIT_TIMEOUT)



# xlsx output


def _write_session_xlsx(all_stats: list[SessionStats], out_path: Path) -> None:
    wb = Workbook()

    # Sheet 1: Summary — one row per session
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = ["Session", "Tenders Seen", "Accepted", "Declined", "Accept Rate", "Total P&L"]
    ws_sum.append(sum_headers)
    for s in all_stats:
        rate = f"{s.tenders_accepted / s.tenders_seen:.1%}" if s.tenders_seen else "0.0%"
        ws_sum.append([s.session, s.tenders_seen, s.tenders_accepted, s.tenders_declined, rate, round(s.total_pnl, 4)])
    # Totals row
    n = len(all_stats)
    if n > 1:
        total_pnl = sum(s.total_pnl for s in all_stats)
        total_seen = sum(s.tenders_seen for s in all_stats)
        total_accepted = sum(s.tenders_accepted for s in all_stats)
        total_declined = sum(s.tenders_declined for s in all_stats)
        overall_rate = f"{total_accepted / total_seen:.1%}" if total_seen else "0.0%"
        ws_sum.append(["TOTAL", total_seen, total_accepted, total_declined, overall_rate, round(total_pnl, 4)])
        # Bold totals row
        for cell in ws_sum[ws_sum.max_row]:
            cell.font = Font(name="Arial", bold=True)
    _style_ws(ws_sum, len(sum_headers))

    # Sheet 2: All tender records across all sessions
    ws_rec = wb.create_sheet("Tender Detail")
    rec_headers = [
        "Session", "Tick", "Tender ID", "Ticker", "Action", "Quantity",
        "Tender Price", "Mid Price", "Edge", "Threshold",
        "Bid Depth Total", "Ask Depth Total", "N Bid Levels", "Best Bid", "Best Ask",
        "Estimated Slippage", "Book Covers Full Qty",
        "Decision", "Unwind Method", "Avg Unwind Price", "P&L", "Residual",
    ]
    ws_rec.append(rec_headers)
    for s in all_stats:
        for r in s.records:
            ws_rec.append([
                s.session, r.tick, r.tender_id, r.ticker, r.action, r.quantity,
                r.tender_price, r.mid_price, r.edge, THRESHOLD,
                r.bid_depth_total, r.ask_depth_total, r.n_bid_levels, r.best_bid, r.best_ask,
                r.estimated_slippage, r.book_covers_full_qty,
                r.decision, r.unwind_method or "", r.fill_price or "", r.pnl or "", r.residual or "",
            ])
    _style_ws(ws_rec, len(rec_headers))
    ws_rec.freeze_panes = "A2"

    # Colour-code decision column
    green_fill = PatternFill("solid", start_color="C6EFCE")
    red_fill = PatternFill("solid", start_color="FFC7CE")
    decision_col = rec_headers.index("Decision") + 1
    for row in ws_rec.iter_rows(min_row=2, min_col=decision_col, max_col=decision_col):
        for cell in row:
            if cell.value == "ACCEPTED":
                cell.fill = green_fill
            elif cell.value == "DECLINED":
                cell.fill = red_fill

    wb.save(out_path)
    wb.close()



# Entry point


def main() -> None:
    client = RotmanSDK(API_KEY=API_KEY, HOST=HOST)
    agent = ReactiveAgent(client, threshold=THRESHOLD)
    all_stats: list[SessionStats] = []
    session = 0

    try:
        while MAX_SESSIONS is None or session < MAX_SESSIONS:
            session += 1
            print(f"\n{'='*60}", flush=True)
            print(f"  SESSION {session}{f'/{MAX_SESSIONS}' if MAX_SESSIONS else ''}  |  threshold={THRESHOLD}", flush=True)
            print(f"{'='*60}", flush=True)

            if not _wait_for_session_start(client):
                print(f"Timed out waiting for session {session}. Stopping.")
                break

            stats = agent.run_session(session)
            all_stats.append(stats)

            print(f"\n[session {session} summary]", flush=True)
            print(f"  Tenders seen:     {stats.tenders_seen}", flush=True)
            print(f"  Accepted:         {stats.tenders_accepted}", flush=True)
            print(f"  Declined:         {stats.tenders_declined}", flush=True)
            print(f"  Total P&L:        {stats.total_pnl:+.4f}", flush=True)

    except KeyboardInterrupt:
        print("\nInterrupted — saving logs.", flush=True)

    if not all_stats:
        print("No sessions completed.")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"reactive_agent_{run_id}.xlsx"
    _write_session_xlsx(all_stats, out_path)
    print(f"\nSaved: {out_path}", flush=True)


if __name__ == "__main__":
    main()